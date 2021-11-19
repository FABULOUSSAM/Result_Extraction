from openpyxl.styles import Alignment  
from openpyxl import Workbook
wb = Workbook()  
ws = wb.active  
ws.cell(row=1, column=1, value="SEAT NO")
ws.cell(row=1, column=2, value="NAME OF STUDENT")
ws.cell(row=1, column=3, value="STATUS")
ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=8)

  
cell = ws.cell(row=1, column=4)  
cell.value = 'Devansh Sharma'  


cell = ws.cell(row=1, column=9)  
cell.value = 'Ddasdasda Sharma'


wb.save('merging.xlsx')  