import cv2
import pytesseract
import numpy as np
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font

from openpyxl.utils import get_column_letter
pytesseract.pytesseract.tesseract_cmd=r'C:\Users\Shubham\AppData\Local\Tesseract-OCR\tesseract.exe'


img = cv2.imread('./Crop_Image/crop_7.png')

gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
blurred = cv2.GaussianBlur(gray, (5, 5), 0)
#edged = cv2.Canny(blurred, 50, 200, 255)
custom_config = r'--oem 3 --psm 6'

val=(pytesseract.image_to_string(blurred, config=custom_config))
values=[]
j=''
for x in val:
    j+=x
    if ' ' in x:
        values.append(j)
        j=""
for x in range(0,len(values)):
    if '\n' in values[x]:
        m=values[x].split('\n')
        values[x]= m[0]
        values.insert(x+1,m[1])

dic={'seat_no':' ','name':' ' ,'status':'','subject':''}
temp=""
nos=[[],[],[]]
sgpi={}
k=0
j=p=q=r=0
list2count=0
list3count=0
list4count=0
list5count=0
storedt=''
for x in range(0,len(values)):
    if x == 0:
        dic['seat_no']=values[x]
    elif values[x].find('(') == -1 and x <= 6 and not(re.search('[0-9]',values[x])):
        temp=temp+' '+(values[x])
        dic['name']=temp
    elif values[x].find('(') != -1 and  x <=9 :
        while values[x].find(')') == -1:
            storedt+=values[x]
            x+=1
        while values[x].find(')') != -1:
            storedt+=values[x]
            dic['subject']=storedt
            x+=1
        k=x-1
    else:
        while j != 10:
            k+=1 
            nos[0].append(values[k])
            j+=1
        list2count=k
        while p!= 10:
            list2count+=1 
            nos[1].append(values[list2count])
            p+=1
            dic['status']=values[list2count+1]
        if(dic['status']=="F"):
            list3count = list2count + 1   
            while q!= 10:
                list3count+=1 
                nos[2].append(values[list3count])
                q+=1
        else:
            list3count = list2count + 1
            while q!= 10:
                list3count+=1 
                if  (re.search("[a-zA-Z]",values[list3count])): 
                    dic['status']+= " " + values[list3count]
                else:
                    nos[2].append(values[list3count])
                    q+=1
        list4count = list3count 
        while True:
            if values[list4count].find(')') != -1  or values[list4count].find('SGPI')!=-1:
                r=list4count
                break
            else:
                list4count+=1 
       
        if (re.search("[a-zA-Z]",values[r-1])) and values[r-1].find('(')!=-1:
            sgpi['result'] =  values[r-4] + ' ' +values[r-3] + ' ' +values[r-2] 
        else:
            sgpi['result'] =  values[r-3] + ' ' +values[r-2] + ' ' +values[r-1] 
        list5count = list4count + 1
        while True:
            list5count+=1
            if values[list5count].find('CGPI') == -1:
                if re.search("[a-zA-Z]",values[list5count]):
                    sgpi[values[list5count]]=""
                elif re.search("[0-9]",values[list5count]):
                    sgpi[values[list5count-1]] = values[list5count]
            else:
                break

#print(values)
# print("dictionary")
print(dic)
# print(nos)
#print(sgpi)
#header1 part

temp_head1=[]
header1={}
out=[]
store1=""
orderCount=0


img3 = cv2.imread('./Header_crop/crop1.png')
gray3 = cv2.cvtColor(img3, cv2.COLOR_BGR2GRAY)
blurred3 = cv2.GaussianBlur(gray3, (5, 5), 0)
#edged = cv2.Canny(blurred, 50, 200, 255)
custom_config = r'--oem 3 --psm 6'
val3=(pytesseract.image_to_string(blurred3, config=custom_config))


for x in range(0,len(val3)):
    store1+=val3[x]
    if ' ' in val3[x] or '\n' in val3[x]:
        temp_head1.append(store1)
        store1=""
for x in range(0,len(temp_head1)):
    if temp_head1[x].find('\n') != -1 :
        m = temp_head1[x].replace('\n','')
        temp_head1[x] = m

for x in range(0,len(temp_head1)):
    if temp_head1[x].find(' ') != -1 :
        m = temp_head1[x].replace(' ','')
        temp_head1[x] = m

for_len=[]
for i in range(0,len(temp_head1)):
    if(temp_head1[i]=="<-----" or temp_head1[i]=="<----"):
        for_len.append(4)
    elif(temp_head1[i]=="<"):
        for_len.append(2)








output=[]
target=['Th','Tw','In','Or','Pr/Or']
target1=['Th(80/32)','Tw(25/10)','In(20/08)','Or(25/10),Pr/Or(25/10)']
for j in range(0,len(temp_head1)):
    if(temp_head1[j] in target1):
        org=temp_head1[j]
        first_part=org[0:2]
        second_part=org[2:len(org)]
        output.append(first_part)
        output.append(second_part)
    else:
        output.append(temp_head1[j])
print(values)
#print(output)
for i in range(0,len(output)):
    if(output[i] in target) :
        out.append(output[i] + output[i+1])
        orderCount=1
    elif re.search("[0-9]",output[i]) and  output[i+2] not in target :
        orderCount = 0 
    elif orderCount != 0  and not(re.search("[0-9]",output[i])) and output[i] not in target:
        out.append('---')
        orderCount=0







# header2 part

temp_head=[]
store=""
header={}
var= ""
z=0

img2 = cv2.imread('./Header_crop/crop2.png')
gray2 = cv2.cvtColor(img2, cv2.COLOR_BGR2GRAY)
blurred2 = cv2.GaussianBlur(gray2, (5, 5), 0)
#edged = cv2.Canny(blurred, 50, 200, 255)
custom_config = r'--oem 3 --psm 6'
val2=(pytesseract.image_to_string(blurred2, config=custom_config))

for x in range(0,len(val2)):
    store+=val2[x]
    if ' ' in val2[x] or '\n' in val2[x]:
        temp_head.append(store)
        store=""
for x in range(0,len(temp_head)):
    if temp_head[x].find('\n') != -1 :
        m = temp_head[x].replace('\n','')
        temp_head[x] = m

for x in temp_head:
    if re.search('[a-zA-Z]',x):
        var+= ' '+x
    if re.search('[0-9]',x) and not(re.search('[a-zA-Z]',x)):
        if len(var) > 0 :
            header[z]=var
            z+=1
            var=""
    if len(var) != 0 :
        header[z] = var


cgpi = ''
cgppi=[]
sgpi['result']+=' '
for x in sgpi['result']:
    cgpi+=x
    if ' ' in cgpi:
        cgppi.append(cgpi)
        cgpi=""
#print(header)







for i in range(0,int(len(out)/2)):
    header1[i]=""
for x in range(0,int(len(out)/2)) :
    header1[x] = out[x % 10]
    header1[x] += '\n ' + out[x+int(len(out)/2)]


#EXCEL PART

wb = Workbook()
ws = wb.active
ws.cell(row=1, column=1, value="SEAT NO")
ws.cell(row=1, column=2, value="NAME OF STUDENT")
ws.cell(row=1, column=3, value="STATUS")

#subject instertion part
start_column=4

for i in range(0,len(header)):
    size=for_len[i]
    value=header[i]
    end_column=start_column+size-1

    ws.merge_cells(start_row=1, end_row=1, start_column=start_column, end_column=end_column);
    cell = ws.cell(row=1, column=start_column)
    cell.value = value
    start_column = end_column+1

#th,tw,in,or iinsertion part
l2,l3=out[0:len(out)//2],out[len(out)//2:]

start_column=4

for i in range(0,len(l2)-2,2):

    cell = ws.cell(row=2, column=start_column)
    cell.value = l2[i]
    
    cell = ws.cell(row=2, column=start_column+1)
    cell.value = l2[i+1]

    cell = ws.cell(row=2, column=start_column+2)
    cell.value = l3[i]

    cell = ws.cell(row=2, column=start_column+3)
    cell.value = l3[i+1]

    start_column+=4

for i in range(len(l2)-2,len(l2)):

    cell = ws.cell(row=2, column=start_column)
    cell.value = l2[i]

    cell = ws.cell(row=2, column=start_column+1)
    cell.value = l3[i]

    start_column+=2

#Pre part insertion

ws.cell(row=3, column=1, value=dic['seat_no'])
ws.cell(row=3, column=2, value=dic['name'])
ws.cell(row=3, column=3, value=dic['status'])


#marks insertion
l2,l3=nos[0],nos[1]
start_column=4

for i in range(0,len(l2)-2,2):

    cell = ws.cell(row=3, column=start_column)
    cell.value = l2[i]
    
    cell = ws.cell(row=3, column=start_column+1)
    cell.value = l2[i+1]

    cell = ws.cell(row=3, column=start_column+2)
    cell.value = l3[i]

    cell = ws.cell(row=3, column=start_column+3)
    cell.value = l3[i+1]

    start_column+=4

for i in range(len(l2)-2,len(l2)):

    cell = ws.cell(row=3, column=start_column)
    cell.value = l2[i]

    cell = ws.cell(row=3, column=start_column+1)
    cell.value = l3[i]

    start_column+=2





#For Text wrap
wrap_alignment = Alignment(wrap_text=True)
for row in ws.iter_rows():
    for cell in row:
       cell.alignment = wrap_alignment

#For coloum width
ws.column_dimensions['B'].width = 30
for i in range(0,len(out)):
    ws.column_dimensions[chr(68+i)].width = 15

#Bold font

for i in range(0,3):
    ws[chr(65+i)+"1"].font = Font(bold=True)
const=0
for i in range(0,len(for_len)):
    ws[chr(68+const)+"1"].font = Font(bold=True)
    const=const+for_len[i]
wb.save("./extract.xlsx")

#print(header1)