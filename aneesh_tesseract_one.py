import pytesseract
import cv2
import numpy as np
import re
import pandas as pd
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
pytesseract.pytesseract.tesseract_cmd=r'C:\Users\Shubham\AppData\Local\Tesseract-OCR\tesseract.exe'

img = cv2.imread('./Crop_Image/crop_6.png')
gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
blurred = cv2.GaussianBlur(gray, (5, 5), 0)
#edged = cv2.Canny(blurred, 50, 200, 255)
custom_config = r'--oem 3 --psm 6'

print(pytesseract.image_to_string(blurred, config=custom_config))

val=(pytesseract.image_to_string(blurred, config=custom_config))
values=[]
j=''
for x in val:
    j+=x
    if ' ' in x:
        values.append(j)
        j=""
for x in range(0,len(values)):
    if '\n' in values[x]:
        m=values[x].split('\n')
        values[x]= m[0]
        values.insert(x+1,m[1])

dic={'seat_no':' ','name':' ' ,'status':''}
temp=""
nos=[[],[],[]]
sgpi={}
k=0
j=p=q=r=0
list2count=0
list3count=0
list4count=0
list5count=0
for x in range(0,len(values)):
    if x == 0:
        dic['seat_no']=values[x]
    elif values[x].find('(') == -1 and x <= 6 and not(re.search('[0-9]',values[x])):
        temp=temp+' '+(values[x])
        dic['name']=temp
    elif values[x].find('(') != -1 and  x <=9 :
        while values[x].find(')') == -1:
            x+=1
        k=x
    else:
        while j != 10:
            k+=1 
            nos[0].append(values[k])
            j+=1
        list2count=k
        while p!= 10:
            list2count+=1 
            nos[1].append(values[list2count])
            p+=1
            dic['status']=values[list2count+1]
        if(dic['status']=="F"):
            list3count = list2count + 1   
            while q!= 10:
                list3count+=1 
                nos[2].append(values[list3count])
                q+=1
        else:
            list3count = list2count + 1
            while q!= 10:
                list3count+=1 
                if  (re.search("[a-zA-Z]",values[list3count])): 
                    dic['status']+= " " + values[list3count]
                else:
                    nos[2].append(values[list3count])
                    q+=1
        list4count = list3count 
        while True:
            if values[list4count].find(')') != -1  or values[list4count].find('SGPI')!=-1:
                r=list4count
                break
            else:
                list4count+=1 
       
        if (re.search("[a-zA-Z]",values[r-1])) and values[r-1].find('(')!=-1:
            sgpi['result'] =  values[r-4] + ' ' +values[r-3] + ' ' +values[r-2] 
        else:
            sgpi['result'] =  values[r-3] + ' ' +values[r-2] + ' ' +values[r-1] 
        list5count = list4count + 1
        while True:
            list5count+=1
            if values[list5count].find('CGPI') == -1:
                if re.search("[a-zA-Z]",values[list5count]):
                    sgpi[values[list5count]]=""
                elif re.search("[0-9]",values[list5count]):
                    sgpi[values[list5count-1]] = values[list5count]
            else:
                break
# print(values)
# print("dictionary")
# print(dic)
# print(nos)
# print(sgpi)



#header1 part
temp_head1=[]
header1={}
out=[]
store1=""
orderCount=0


img3 = cv2.imread('./Header_crop/crop2.png')
gray3 = cv2.cvtColor(img3, cv2.COLOR_BGR2GRAY)
blurred3 = cv2.GaussianBlur(gray3, (5, 5), 0)
#edged = cv2.Canny(blurred, 50, 200, 255)
custom_config = r'--oem 3 --psm 6'
val3=(pytesseract.image_to_string(blurred3, config=custom_config))


for x in range(0,len(val3)):
    store1+=val3[x]
    if ' ' in val3[x] or '\n' in val3[x]:
        temp_head1.append(store1)
        store1=" "

for x in range(0,len(temp_head1)):
    if temp_head1[x].find('\n') != -1 :
        m = temp_head1[x].replace('\n','')
        temp_head1[x] = m

for x in range(0,len(temp_head1)):
    if temp_head1[x].find(' ') != -1 :
        m = temp_head1[x].replace(' ','')
        temp_head1[x] = m

target=['Th','Tw','In','Or']
target1=['Th(80/32)','Tw(25/10)','In(20/08)','Or(25/10)']
for j in range(0,len(temp_head1)):
    if(temp_head1[j] in target):
        out.append(temp_head1[j])
        out.append(temp_head1[j+1])
        orderCount=1
    elif(temp_head1[j] in target1):
        org=temp_head1[j]
        first_part=org[0:2]
        second_part=org[2:len(org)]
        out.append(first_part)
        out.append(second_part)
        orderCount=1
    elif re.search("[0-9]",temp_head1[x]) and  temp_head1[x+2] not in target :
        orderCount = 0 
    elif orderCount != 0  and not(re.search("[0-9]",temp_head1[x])) and temp_head1[x] not in target:
        out.append('---')
        orderCount=0
print(out)
for i in range(0,int(len(out)/2)):
    header1[i]=""
for x in range(0,int(len(out)/2)) :
    header1[x] = out[x % 10]
    header1[x] += '\n ' + out[x+int(len(out)/2)]


# header2 part

temp_head=[]
store=""
header={}
var= ""
z=0

img2 = cv2.imread('./Header_crop/crop3.png')
gray2 = cv2.cvtColor(img2, cv2.COLOR_BGR2GRAY)
blurred2 = cv2.GaussianBlur(gray2, (5, 5), 0)
#edged = cv2.Canny(blurred, 50, 200, 255)
custom_config = r'--oem 3 --psm 6'
val2=(pytesseract.image_to_string(blurred2, config=custom_config))

for x in range(0,len(val2)):
    store+=val2[x]
    if ' ' in val2[x] or '\n' in val2[x]:
        temp_head.append(store)
        store=""
for x in range(0,len(temp_head)):
    if temp_head[x].find('\n') != -1 :
        m = temp_head[x].replace('\n','')
        temp_head[x] = m

for x in temp_head:
    if re.search('[a-zA-Z]',x):
        var+= ' '+x
    if re.search('[0-9]',x) and not(re.search('[a-zA-Z]',x)):
        if len(var) > 0 :
            header[z]=var
            z+=1
            var=""
    if len(var) != 0 :
        header[z] = var


cgpi = ''
cgppi=[]
sgpi['result']+=' '
for x in sgpi['result']:
    cgpi+=x
    if ' ' in cgpi:
        cgppi.append(cgpi)
        cgpi=""



# for x in range(0,len(nos)):
#     for i in range(0,len(nos[x])):
#         print(nos[x][i])
        # if not(re.search('[0-9]',x[i])):
        #     nos.insert(i,'---')


print(dic)
print(sgpi)
print(nos)

print(header1)
#Excel Part


ws['C1'] = 'Seat No'
ws['D1'] = 'Name'
ws['E1'] = 'Status'
ws['F1'] = 'CGPI'

ws['G1'] = header[0]
ws['H1'] = header[1]
ws['I1'] = header[2]
ws['J1'] = header[3]
ws['K1'] = header[4]
ws['L1'] = header[5]

ws['G2'] = header1[0]+header1[1]
ws['H2'] = header1[2]+header1[3]
ws['I2'] = header1[4]+header1[5]
ws['J2'] = header1[6]+header1[7]
ws['K2'] = header1[8]
ws['L2'] = header1[9]



ws.column_dimensions['D'].width = 30
ws.column_dimensions['G'].width = 35
ws.column_dimensions['H'].width = 35
ws.column_dimensions['I'].width = 35
ws.column_dimensions['J'].width = 35
ws.column_dimensions['K'].width = 35
ws.column_dimensions['L'].width = 35




params1 = 'CDE'
params2 ='GHIJKL'  
counters=3
lcounter=3
noscounter=0
value2=""

for i in params1:
    if i ==  'C':
        ws[i+str(counters)] =dic['seat_no']
    if i== "D":
        ws[i+str(counters)]=dic['name']
    if i == "E":
        ws[i+str(counters)]=dic['status']




while noscounter != len(nos[0]):
    for i in params2:
        if i != 'K' and i != 'L':   
            for x in range(0,len(nos)):
                value2+=nos[x][noscounter] + nos[x][noscounter+1]
                ws[i+str(lcounter)]=value2
            value2=""
            noscounter+=2 
           
        else:
            for x in range(0,len(nos)):
                value2+=nos[x][noscounter]
                ws[i+str(lcounter)] =value2
            value2=""
            noscounter+=1 

        
wb.save("./newextract.xlsx")

print(sgpi)

